#!/usr/bin/env python3
"""Parse Leapfrog Hopper Safety Grade XLSX → JSON on stdout."""

import json
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write("openpyxl required: pip install openpyxl\n")
    sys.exit(1)


def normalize_ccn(raw):
    digits = "".join(c for c in str(raw or "") if c.isdigit())
    return digits.zfill(6)[-6:] if digits else ""


def parse_grade(raw):
    if raw is None:
        return None
    s = str(raw).strip().upper()
    if s in ("A", "B", "C", "D", "F"):
        return s
    return None


def main():
    if len(sys.argv) < 2:
        sys.stderr.write("Usage: ingestLeapfrog.py file.xlsx\n")
        sys.exit(1)

    path = sys.argv[1]
    wb = load_workbook(path, read_only=True, data_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if name.lower().startswith("spring") or name.lower().startswith("fall"):
            sheet_name = name
            break
    if not sheet_name:
        for name in wb.sheetnames:
            if "spring" in name.lower() or "fall" in name.lower():
                sheet_name = name
                break
    if not sheet_name:
        sys.stderr.write(f"No grade sheet found in {wb.sheetnames}\n")
        sys.exit(1)

    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows)]

    def col(*names):
        for n in names:
            for i, h in enumerate(headers):
                if h.lower() == n.lower():
                    return i
        return None

    ccn_col = col("CMS_Certification_Number", "CMS Certification Number")
    grade_col = col("Hospital Grade", "Grade")
    score_col = col("Hospital Score", "Score")
    name_col = col("Facility_Name", "Hospital Name")

    if ccn_col is None or grade_col is None:
        sys.stderr.write(f"Missing CMS/Grade columns in {sheet_name}: {headers[:20]}\n")
        sys.exit(1)

    release = sheet_name.replace("20", " 20").replace("Spring", "Spring ").replace("Fall", "Fall ")
    grades = {}

    for row in rows:
        if not row or not row[ccn_col]:
            continue
        ccn = normalize_ccn(row[ccn_col])
        if not ccn or ccn.startswith("XX"):
            continue
        grade = parse_grade(row[grade_col])
        score = row[score_col] if score_col is not None else None
        try:
            score_val = float(score) if score is not None else None
        except (TypeError, ValueError):
            score_val = None

        leapfrog_id = f"{ccn[:2]}-{ccn[2:]}"
        status = "graded" if grade else "not_assigned"
        grades[ccn] = {
            "facilityId": ccn,
            "grade": grade,
            "score": score_val,
            "status": status,
            "release": release.strip(),
            "profileUrl": f"https://www.hospitalsafetygrade.org/h/{leapfrog_id}",
        }

    json.dump(grades, sys.stdout)
    wb.close()


if __name__ == "__main__":
    main()
